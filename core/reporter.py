"""
XLSX test-execution reporter for the Flitesports automation framework.

Each automation module instantiates one ``XlsxReporter``, calls
``add_step()`` after every UI action, and calls ``save()`` in a
``finally`` block so the report is always written regardless of outcome.

Output layout
─────────────
  Sheet "Summary"
      Module name, run date, start/end time, total duration (seconds),
      step counts, and overall PASSED / FAILED status.

  Sheet "Steps"
      #  |  Step Name  |  Status  |  Detail / Notes  |  Timestamp
      |  Step Duration (s)  |  Cumulative Time (s)

The report filename pattern is:
    ``<ModuleName>_<YYYYMMDD_HHMMSS>.xlsx``
"""

import datetime
import logging
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import REPORTS_DIR

log = logging.getLogger(__name__)

# ── Colour tokens ──────────────────────────────────────────────────────────────
_C = {
    "navy":       "1A1A2E",
    "deep_blue":  "0F3460",
    "mid_blue":   "16213E",
    "white":      "FFFFFF",
    "pass_bg":    "E6F9ED",
    "pass_fg":    "28A745",
    "fail_bg":    "FDECEA",
    "fail_fg":    "DC3545",
    "info_bg":    "E8F0FE",
    "info_fg":    "1A73E8",
    "row_alt":    "F4F7FF",
    "row_white":  "FFFFFF",
    "border":     "D0D7E5",
    "summary_bg": "D6E4F0",
}

StatusType = Literal["PASS", "FAIL", "INFO"]


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(
    bold: bool = False,
    color: str = "1A1A2E",
    size: int = 10,
    italic: bool = False,
) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)


def _border() -> Border:
    side = Side(style="thin", color=_C["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


# ── Reporter class ─────────────────────────────────────────────────────────────

class XlsxReporter:
    """
    Collects test-step results during a module run and writes a two-sheet
    XLSX report when ``save()`` is called.
    """

    def __init__(self, module_name: str) -> None:
        """
        Args:
            module_name: Human-readable name displayed in the report header
                         and used to derive the output filename.
        """
        self.module_name     = module_name
        self._steps: list[dict] = []
        self._start_time     = datetime.datetime.now()
        self._last_step_time = self._start_time

    # ── Public API ─────────────────────────────────────────────────────────────

    def add_step(
        self,
        name: str,
        status: StatusType,
        detail: str = "",
    ) -> None:
        """
        Record the outcome of a completed test step.

        Computes two timing values:
          * **step_secs**    – elapsed seconds since the previous ``add_step`` call
                               (or since construction for the first step).
          * **cumulative_s** – total elapsed seconds from module start.

        Args:
            name:   Short, human-readable description of the step (≤ 80 chars).
            status: ``'PASS'``, ``'FAIL'``, or ``'INFO'``.
            detail: Supplementary data such as generated values, URLs, or
                    error messages.  May be empty.
        """
        now           = datetime.datetime.now()
        step_duration = round((now - self._last_step_time).total_seconds(), 2)
        cumulative    = round((now - self._start_time).total_seconds(), 2)

        self._steps.append(
            {
                "name":         name,
                "status":       status.upper(),
                "detail":       detail,
                "timestamp":    now.strftime("%H:%M:%S"),
                "step_secs":    step_duration,
                "cumulative_s": cumulative,
            }
        )
        self._last_step_time = now

        # Mirror to the console log
        level = logging.ERROR if status.upper() == "FAIL" else logging.INFO
        log.log(
            level,
            "[%-4s]  %-55s  %s",
            status.upper(),
            name,
            f"→ {detail}" if detail else "",
        )

    def save(self) -> Path:
        """
        Finalise the report, write it to ``REPORTS_DIR``, and return the path.

        Always call this method inside a ``finally`` block so the report is
        written even when the module raises an exception.

        Returns:
            The absolute ``Path`` of the saved ``.xlsx`` file.
        """
        now      = datetime.datetime.now()
        duration = round((now - self._start_time).total_seconds(), 2)
        total    = len(self._steps)
        passed   = sum(1 for s in self._steps if s["status"] == "PASS")
        failed   = sum(1 for s in self._steps if s["status"] == "FAIL")
        info     = sum(1 for s in self._steps if s["status"] == "INFO")
        overall  = "PASSED" if failed == 0 else "FAILED"

        safe_name = self.module_name.replace(" ", "_").replace("/", "-")
        timestamp = self._start_time.strftime("%Y%m%d_%H%M%S")
        path      = REPORTS_DIR / f"{safe_name}_{timestamp}.xlsx"

        wb = Workbook()
        self._write_summary_sheet(wb, now, duration, total, passed, failed, info, overall)
        self._write_steps_sheet(wb)
        wb.save(str(path))

        log.info("XLSX report saved → %s", path)
        return path

    # ── Private helpers ────────────────────────────────────────────────────────

    def _write_summary_sheet(
        self,
        wb: Workbook,
        end_time: datetime.datetime,
        duration: float,
        total: int,
        passed: int,
        failed: int,
        info: int,
        overall: str,
    ) -> None:
        ws          = wb.active
        ws.title    = "Summary"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 50

        # ── Title banner ──────────────────────────────────────────────────────
        ws.merge_cells("A1:B1")
        title_cell           = ws["A1"]
        title_cell.value     = f"Flitesports Automation  —  {self.module_name}"
        title_cell.font      = _font(bold=True, color=_C["white"], size=13)
        title_cell.fill      = _fill(_C["navy"])
        title_cell.alignment = _align("center")
        ws.row_dimensions[1].height = 34

        # ── Key-value summary rows ────────────────────────────────────────────
        rows_data = [
            ("Module",           self.module_name,                               False, "1A1A2E"),
            ("Run Date",         self._start_time.strftime("%d %B %Y"),          False, "1A1A2E"),
            ("Start Time",       self._start_time.strftime("%H:%M:%S"),          False, "1A1A2E"),
            ("End Time",         end_time.strftime("%H:%M:%S"),                  False, "1A1A2E"),
            ("Total Duration",   f"{duration:.2f} seconds",                      True,  _C["deep_blue"]),
            ("Total Steps",      str(total),                                     False, "1A1A2E"),
            ("Passed",           str(passed),                                    True,  _C["pass_fg"]),
            ("Failed",           str(failed),                                    True,  _C["fail_fg"] if failed else "1A1A2E"),
            ("Info / Recorded",  str(info),                                      False, "1A1A2E"),
            ("Overall Result",   overall,                                        True,  _C["pass_fg"] if overall == "PASSED" else _C["fail_fg"]),
            ("Project",          "Flitesports – Shopify & eCommerce Automation", False, "1A1A2E"),
            ("Author",           "Vivek Jain – QA Automation Engineer",          False, "1A1A2E"),
        ]

        for label, value, bold_val, val_color in rows_data:
            r = ws.max_row + 1
            ws.row_dimensions[r].height = 22

            lc            = ws.cell(r, 1, label)
            lc.font       = _font(bold=True, color=_C["deep_blue"])
            lc.fill       = _fill(_C["summary_bg"])
            lc.border     = _border()
            lc.alignment  = _align("left")

            vc            = ws.cell(r, 2, value)
            vc.font       = _font(bold=bold_val, color=val_color)
            vc.border     = _border()
            vc.alignment  = _align("left")

    def _write_steps_sheet(self, wb: Workbook) -> None:
        ws          = wb.create_sheet("Steps")
        ws.sheet_view.showGridLines = False

        headers    = [
            "#",
            "Step Name",
            "Status",
            "Detail / Notes",
            "Timestamp",
            "Step Duration (s)",
            "Cumulative Time (s)",
        ]
        col_widths = [5, 54, 10, 56, 12, 20, 22]

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Header row ────────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 26
        for col_idx, header in enumerate(headers, 1):
            cell           = ws.cell(1, col_idx, header)
            cell.font      = _font(bold=True, color=_C["white"], size=10)
            cell.fill      = _fill(_C["navy"])
            cell.border    = _border()
            cell.alignment = _align("center")

        # ── Data rows ─────────────────────────────────────────────────────────
        _status_styles = {
            "PASS": (_C["pass_bg"], _C["pass_fg"]),
            "FAIL": (_C["fail_bg"], _C["fail_fg"]),
            "INFO": (_C["info_bg"], _C["info_fg"]),
        }

        for row_num, step in enumerate(self._steps, 2):
            status      = step["status"]
            s_bg, s_fg  = _status_styles.get(status, (_C["row_white"], "1A1A2E"))
            row_bg      = _C["row_alt"] if row_num % 2 == 0 else _C["row_white"]

            ws.row_dimensions[row_num].height = 20

            values = [
                row_num - 1,
                step["name"],
                status,
                step["detail"],
                step["timestamp"],
                step["step_secs"],
                step["cumulative_s"],
            ]
            # Columns that use centre-alignment
            centre_cols = {1, 3, 5, 6, 7}

            for col_idx, val in enumerate(values, 1):
                cell           = ws.cell(row_num, col_idx, val)
                cell.border    = _border()
                cell.alignment = _align(
                    "center" if col_idx in centre_cols else "left",
                    wrap=col_idx in (2, 4),
                )

                if col_idx == 3:  # Status badge column
                    cell.font = _font(bold=True, color=s_fg)
                    cell.fill = _fill(s_bg)
                else:
                    cell.font = _font(color="333333")
                    cell.fill = _fill(row_bg)

        # Freeze the header row so it stays visible while scrolling
        ws.freeze_panes = "A2"
