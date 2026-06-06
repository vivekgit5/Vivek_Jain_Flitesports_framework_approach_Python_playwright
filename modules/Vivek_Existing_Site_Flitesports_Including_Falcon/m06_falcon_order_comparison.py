"""
Module M06 – Falcon Stage: Order Exporter vs Order Generator Comparison
========================================================================
Automates a three-phase end-to-end verification workflow on the Falcon
staging environment.

Phase 1 – Data Acquisition (browser automation)
    • Navigate to the Falcon App Launcher.
    • Launch the Order Exporter, configure it for order numbers 1050-1070
      (all channels), and download the resulting Excel workbook.
    • Launch the Order Generator (Production Order tool), configure the same
      order range, and download the resulting Excel workbook.

Phase 2 – Data Comparison (offline)
    • Parse both Excel workbooks tab by tab.
    • Skip excluded tabs: "Island of Misfit Toys", "Unknown Programs".
    • For every remaining tab, align rows by Order ID and compare every
      field between the two sources.

Phase 3 – Reporting
    • Falcon_Stage_Order_Comparison_Report_<timestamp>.xlsx
        Full field-by-field comparison table (matched rows green,
        mismatched cells red).
    • Falcon_Stage_Mismatches_Report_<timestamp>.xlsx
        Only orders/fields that diverge; each mismatch is self-described
        with Order ID, source tab, and both values.
    • Standard XlsxReporter execution report (run timing, step results).

Usage
-----
    python -m modules.m06_falcon_order_comparison

Requirements (in addition to base framework dependencies)
----------------------------------------------------------
    pip install openpyxl pandas
"""

import datetime
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import Download, Page, sync_playwright

from config.settings import (
    DOWNLOADS_DIR,
    FALCON_APP_LAUNCHER_URL,
    FALCON_EXPORT_CHANNEL,
    FALCON_ORDER_RANGE,
    FALCON_PRODUCTION_ORDER,
    REPORTS_DIR,
    VIEWPORT,
)
from core.reporter import XlsxReporter

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

MODULE_NAME = "M06 – Falcon Order Exporter vs Generator Comparison"

# ── Excluded tabs (case-insensitive) ──────────────────────────────────────────
EXCLUDED_TABS: frozenset = frozenset({
    "island of misfit toys",
    "unknown programs",
    "unknown program",
})

# ── Column-header candidate sets ──────────────────────────────────────────────
ORDER_ID_HEADER_CANDIDATES: frozenset = frozenset({
    "order id", "order #", "order#", "order number", "ordernumber",
    "order_id", "order no", "#", "shopify order id", "order",
    "order num", "order-id", "order ref", "order no.", "order#.",
})
PLAYER_NAME_CANDIDATES: frozenset = frozenset({
    "player name", "name", "player", "athlete name", "athlete",
    "registered player name", "customer name", "full name",
})
PLAYER_NUMBER_CANDIDATES: frozenset = frozenset({
    "player number", "number", "player #", "player#", "jersey number",
    "player no", "jersey no", "no", "num",
})


# ══════════════════════════════════════════════════════════════════════════════
# §1 – Data models
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class FieldResult:
    """Comparison outcome for a single field within one order row."""
    column:          str
    exporter_value:  Any
    generator_value: Any

    @property
    def matches(self) -> bool:
        return _normalise(self.exporter_value) == _normalise(self.generator_value)


@dataclass
class OrderResult:
    """Aggregated comparison for one Order ID within one Excel sheet/tab."""
    order_id: str
    sheet:    str
    fields:   List[FieldResult] = field(default_factory=list)

    @property
    def has_mismatch(self) -> bool:
        return any(not f.matches for f in self.fields)

    @property
    def mismatch_count(self) -> int:
        return sum(1 for f in self.fields if not f.matches)

    @property
    def match_count(self) -> int:
        return sum(1 for f in self.fields if f.matches)


# ══════════════════════════════════════════════════════════════════════════════
# §2 – Utility helpers
# ══════════════════════════════════════════════════════════════════════════════

def _normalise(value: Any) -> str:
    """Reduce a cell value to a stripped, lowercase, whitespace-normalised string."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _safe_str(value: Any, fallback: str = "—") -> str:
    """Convert a cell value to a printable string; replace blank/NaN with fallback."""
    if value is None:
        return fallback
    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass
    result = str(value).strip()
    return result if result else fallback


# ══════════════════════════════════════════════════════════════════════════════
# §3 – Excel parsing
# ══════════════════════════════════════════════════════════════════════════════

def _detect_column(df: pd.DataFrame, candidates: frozenset) -> Optional[str]:
    """Return the first column whose lowercased header is in *candidates*, or None."""
    for col in df.columns:
        if str(col).strip().lower() in candidates:
            return col
    return None


def _detect_order_id_column(df: pd.DataFrame) -> Optional[str]:
    """Scan DataFrame columns for a recognised Order ID header. Returns col name or None."""
    for col in df.columns:
        if str(col).strip().lower() in ORDER_ID_HEADER_CANDIDATES:
            return col
    return None


def _find_header_row(path: Path, sheet_name: str, max_search: int = 6) -> int:
    """
    Scan the first *max_search* rows for the real header row.

    Some Generator sheets begin with a merged-cell title row that produces
    ``Unnamed`` column headers.  This function skips those rows and returns
    the index of the first row with at least 3 named columns.
    """
    for row_idx in range(max_search):
        try:
            sample = pd.read_excel(
                path, sheet_name=sheet_name,
                header=row_idx, nrows=0, dtype=str,
            )
            named = [c for c in sample.columns if not str(c).startswith("Unnamed")]
            if len(named) >= 3:
                return row_idx
        except Exception:
            break
    return 0


def parse_excel_workbook(path: Path) -> Dict[str, pd.DataFrame]:
    """
    Load every non-excluded sheet from *path* into a dict of
    ``{sheet_name: DataFrame}``.

    * Auto-detects merged-cell header rows.
    * Drops fully-empty rows and columns.
    * Drops residual ``Unnamed`` columns.
    """
    log.info("Parsing workbook: %s", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: Dict[str, pd.DataFrame] = {}

    for name in wb.sheetnames:
        if name.strip().lower() in EXCLUDED_TABS:
            log.info("  Excluded tab skipped: '%s'", name)
            continue
        log.info("  Reading tab: '%s'", name)
        try:
            header_row = _find_header_row(path, name)
            df = pd.read_excel(path, sheet_name=name, header=header_row, dtype=str)
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
            df.reset_index(drop=True, inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            if df.empty or len(df.columns) < 2:
                log.info("    Tab '%s' has no usable data — skipped.", name)
                continue
            sheets[name] = df
        except Exception as exc:
            log.warning("  Could not parse tab '%s': %s", name, exc)

    wb.close()
    return sheets


# ══════════════════════════════════════════════════════════════════════════════
# §4 – Comparison engine
# ══════════════════════════════════════════════════════════════════════════════

def _flatten_sheets(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Concatenate all sheets into one DataFrame, tagging each row with its source tab."""
    frames: List[pd.DataFrame] = []
    for sheet_name, df in sheets.items():
        tagged = df.copy()
        tagged["__sheet__"] = sheet_name
        frames.append(tagged)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


def _sort_key(order_id: str) -> tuple:
    """Natural sort key: numeric part first, then raw string."""
    digits = re.sub(r"\D", "", order_id)
    return (int(digits) if digits else 0, order_id)


def _parse_sheet_label(label: str):
    """
    Decompose a composite sheet-label into (exporter_tab, generator_tab, line_id).

    Example input:
        "Exporter [Sublimation]  →  Generator [PRICING]  |  Player Name: vivek"
    """
    exp_m   = re.search(r"Exporter \[([^\]]+)\]", label)
    gen_m   = re.search(r"Generator \[([^\]]+)\]", label)
    line_id = label.split("|", 1)[1].strip() if "|" in label else "Line 1"
    return (
        exp_m.group(1) if exp_m else "—",
        gen_m.group(1) if gen_m else "—",
        line_id,
    )


def compare_workbooks(
    exporter_sheets: Dict[str, pd.DataFrame],
    generator_sheets: Dict[str, pd.DataFrame],
) -> List[OrderResult]:
    """
    Compare both workbooks by Order ID across all eligible tabs.

    Because the Order Exporter groups rows by product type while the Order
    Generator groups rows by team name, tab names do not align.  Both datasets
    are therefore flattened into single frames before matching by Order ID.

    Row-matching within a given Order ID:
      1. By player name (if both sources have a player-name column).
      2. By player/jersey number (fallback).
      3. Positionally (last resort).

    Returns a flat list of OrderResult objects (one entry per line item).
    """
    exp_flat = _flatten_sheets(exporter_sheets)
    gen_flat = _flatten_sheets(generator_sheets)

    if exp_flat.empty or gen_flat.empty:
        log.warning("One or both workbooks yielded no usable data.")
        return []

    exp_id_col = _detect_order_id_column(exp_flat)
    gen_id_col = _detect_order_id_column(gen_flat)

    if not exp_id_col or not gen_id_col:
        log.warning("Order-ID column not found in one or both workbooks.")
        return []

    exp_flat = exp_flat.copy()
    gen_flat = gen_flat.copy()
    exp_flat["__norm_id__"] = exp_flat[exp_id_col].apply(_normalise)
    gen_flat["__norm_id__"] = gen_flat[gen_id_col].apply(_normalise)

    all_ids = sorted(
        (set(exp_flat["__norm_id__"].unique()) | set(gen_flat["__norm_id__"].unique())) - {""},
        key=_sort_key,
    )

    _COL_ALIASES: Dict[str, str] = {
        "player#":                "PLAYER NUMBER",
        "player #":               "PLAYER NUMBER",
        "registered player name": "PLAYER NAME",
        "jersey name":            "JERSEY NAME",
    }
    _internal = {"__sheet__", "__norm_id__", exp_id_col, gen_id_col}

    def _canonical(col: str) -> str:
        return _COL_ALIASES.get(col.strip().lower(), col.strip().upper())

    exp_canon: Dict[str, str] = {_canonical(c): c for c in exp_flat.columns if c not in _internal}
    gen_canon: Dict[str, str] = {_canonical(c): c for c in gen_flat.columns if c not in _internal}
    common_canonical = sorted(set(exp_canon) & set(gen_canon))
    compare_triples  = [(cn, exp_canon[cn], gen_canon[cn]) for cn in common_canonical]

    log.info(
        "Flattened: %d exporter rows / %d generator rows  |  "
        "%d unique Order IDs  |  %d comparison columns",
        len(exp_flat), len(gen_flat), len(all_ids), len(compare_triples),
    )

    results: List[OrderResult] = []

    for oid in all_ids:
        exp_rows = exp_flat[exp_flat["__norm_id__"] == oid]
        gen_rows = gen_flat[gen_flat["__norm_id__"] == oid]

        exp_tab = ", ".join(exp_rows["__sheet__"].unique().tolist()) if not exp_rows.empty else "—"
        gen_tab = ", ".join(gen_rows["__sheet__"].unique().tolist()) if not gen_rows.empty else "—"
        sheet_label = f"Exporter [{exp_tab}]  →  Generator [{gen_tab}]"

        exp_name_col   = _detect_column(exp_rows, PLAYER_NAME_CANDIDATES)
        gen_name_col   = _detect_column(gen_rows, PLAYER_NAME_CANDIDATES)
        exp_number_col = _detect_column(exp_rows, PLAYER_NUMBER_CANDIDATES)
        gen_number_col = _detect_column(gen_rows, PLAYER_NUMBER_CANDIDATES)

        if exp_name_col and gen_name_col and not exp_rows.empty and not gen_rows.empty:
            match_key, exp_key_col, gen_key_col = "player name", exp_name_col, gen_name_col
        elif exp_number_col and gen_number_col and not exp_rows.empty and not gen_rows.empty:
            match_key, exp_key_col, gen_key_col = "player number", exp_number_col, gen_number_col
        else:
            match_key = exp_key_col = gen_key_col = None

        def _build(exp_row: Any, gen_row: Any, label: str) -> OrderResult:
            r = OrderResult(order_id=oid, sheet=label)
            for (cn, ec, gc) in compare_triples:
                r.fields.append(FieldResult(
                    cn,
                    exp_row.get(ec) if exp_row is not None else None,
                    gen_row.get(gc) if gen_row is not None else None,
                ))
            return r

        if match_key:
            exp_by_key = {_normalise(r[exp_key_col]): r for _, r in exp_rows.iterrows()}
            gen_by_key = {_normalise(r[gen_key_col]): r for _, r in gen_rows.iterrows()}
            for key in sorted(set(exp_by_key) | set(gen_by_key)):
                lbl = f"{sheet_label}  |  {match_key.title()}: {key or 'unknown'}"
                results.append(_build(exp_by_key.get(key), gen_by_key.get(key), lbl))
        else:
            exp_list = [r for _, r in exp_rows.iterrows()]
            gen_list = [r for _, r in gen_rows.iterrows()]
            for i in range(max(len(exp_list), len(gen_list), 1)):
                exp_row = exp_list[i] if i < len(exp_list) else None
                gen_row = gen_list[i] if i < len(gen_list) else None
                results.append(_build(exp_row, gen_row, f"{sheet_label}  |  Line {i + 1}"))

    log.info("Comparison complete: %d line-item result(s).", len(results))
    return results


# ══════════════════════════════════════════════════════════════════════════════
# §5 – XLSX report generation (Comparison + Mismatches)
# ══════════════════════════════════════════════════════════════════════════════

# ── Colour palette ─────────────────────────────────────────────────────────────
_F_HDR_CMP    = PatternFill("solid", fgColor="0F3460")
_F_HDR_MM     = PatternFill("solid", fgColor="7B1E1E")
_F_ORD_OK     = PatternFill("solid", fgColor="1A5E3A")
_F_ORD_BAD    = PatternFill("solid", fgColor="8B1A1A")
_F_ROW_OK     = PatternFill("solid", fgColor="F0FFF4")
_F_ROW_BAD    = PatternFill("solid", fgColor="FFF0F0")
_F_CELL_BAD   = PatternFill("solid", fgColor="FDECEA")
_F_STAT_TOT   = PatternFill("solid", fgColor="E8F0FE")
_F_STAT_OK    = PatternFill("solid", fgColor="E6F9ED")
_F_STAT_BAD   = PatternFill("solid", fgColor="FDECEA")

_FT_TITLE     = Font(name="Calibri", bold=True,  color="FFFFFF", size=13)
_FT_META      = Font(name="Calibri", italic=True, color="CCCCCC", size=9)
_FT_HDR       = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
_FT_ORD       = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
_FT_OK        = Font(name="Calibri", bold=True,  color="1A7A3C", size=10)
_FT_BAD       = Font(name="Calibri", bold=True,  color="C0392B", size=10)
_FT_NORM      = Font(name="Calibri",              size=10)
_FT_NORMB     = Font(name="Calibri", bold=True,   size=10)
_FT_SRC       = Font(name="Calibri", italic=True, color="444444", size=9)
_FT_STAT_V    = Font(name="Calibri", bold=True,   size=14)
_FT_STAT_L    = Font(name="Calibri", color="666666", size=9)

_AL_C  = Alignment(horizontal="center", vertical="center")
_AL_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_BORD  = Border(
    left=Side(style="thin", color="D0D4E0"),
    right=Side(style="thin", color="D0D4E0"),
    top=Side(style="thin", color="D0D4E0"),
    bottom=Side(style="thin", color="D0D4E0"),
)


def _sc(cell, fill=None, font=None, align=None) -> None:
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    cell.border = _BORD


def _title_row(ws, title: str, n: int, fill, meta: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(1, 1, title)
    c.fill = fill; c.font = _FT_TITLE; c.alignment = _AL_C
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(2, 1, meta)
    c.fill = PatternFill("solid", fgColor="16213E"); c.font = _FT_META; c.alignment = _AL_C
    ws.row_dimensions[2].height = 15


def _stat_block(ws, stats: list, start_row: int) -> int:
    for col, (lbl, val, fill) in enumerate(stats, 1):
        lc = ws.cell(start_row,     col, lbl)
        vc = ws.cell(start_row + 1, col, val)
        _sc(lc, fill=fill, font=_FT_STAT_L, align=_AL_C)
        _sc(vc, fill=fill, font=_FT_STAT_V, align=_AL_C)
    ws.row_dimensions[start_row].height     = 16
    ws.row_dimensions[start_row + 1].height = 32
    return start_row + 2


def _hdr_row(ws, hdrs: list, row: int, fill) -> None:
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row, col, h)
        _sc(c, fill=fill, font=_FT_HDR, align=_AL_C)
    ws.row_dimensions[row].height = 20


def _build_summary_sheet(ws, results: List[OrderResult], ts: str, duration: int) -> None:
    """Write the executive-summary sheet shared by both report workbooks."""
    unique_oids     = len({r.order_id for r in results})
    total           = len(results)
    orders_ok       = sum(1 for r in results if not r.has_mismatch)
    orders_bad      = sum(1 for r in results if r.has_mismatch)
    fields_match    = sum(r.match_count    for r in results)
    fields_mismatch = sum(r.mismatch_count for r in results)
    N               = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    c = ws.cell(1, 1, "Falcon Stage — Order Exporter vs Generator: Executive Summary")
    c.fill = _F_HDR_CMP; c.font = _FT_TITLE; c.alignment = _AL_C
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
    c = ws.cell(2, 1, (
        f"Order Range: {FALCON_ORDER_RANGE}  |  Channel: {FALCON_EXPORT_CHANNEL}  |  "
        f"Generated: {ts}  |  Run Duration: {duration}s"
    ))
    c.fill = PatternFill("solid", fgColor="16213E"); c.font = _FT_META; c.alignment = _AL_C
    ws.row_dimensions[2].height = 15

    kpis = [
        ("Unique Orders",       unique_oids,     _F_STAT_TOT, "0F3460"),
        ("Line Items Compared", total,           _F_STAT_TOT, "0F3460"),
        ("Fully Matched",       orders_ok,       _F_STAT_OK,  "1A7A3C"),
        ("Orders with Issues",  orders_bad,      _F_STAT_BAD, "C0392B"),
        ("Fields Matched",      fields_match,    _F_STAT_OK,  "1A7A3C"),
        ("Fields Mismatched",   fields_mismatch, _F_STAT_BAD, "C0392B"),
    ]
    for col, (lbl, val, fill, color) in enumerate(kpis, 1):
        cv = ws.cell(4, col, val)
        cv.fill = fill; cv.font = Font(name="Calibri", bold=True, color=color, size=20)
        cv.alignment = _AL_C; cv.border = _BORD
        cl = ws.cell(5, col, lbl)
        cl.fill = fill; cl.font = _FT_STAT_L; cl.alignment = _AL_C; cl.border = _BORD
    ws.row_dimensions[4].height = 38
    ws.row_dimensions[5].height = 16

    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=N)
    c = ws.cell(7, 1, "  Order-by-Order Breakdown")
    c.fill = PatternFill("solid", fgColor="2C3E50")
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10); c.alignment = _AL_L
    ws.row_dimensions[7].height = 18

    bdrs = ["Order ID", "Line Items", "Fields Compared", "Matched Fields", "Mismatched Fields", "Status"]
    for col, h in enumerate(bdrs, 1):
        c = ws.cell(8, col, h)
        _sc(c, fill=_F_HDR_CMP, font=_FT_HDR, align=_AL_C)
    ws.row_dimensions[8].height = 18

    by_order: Dict[str, List[OrderResult]] = {}
    for r in results:
        by_order.setdefault(r.order_id, []).append(r)

    for row_idx, oid in enumerate(sorted(by_order.keys(), key=_sort_key), 9):
        items   = by_order[oid]
        n_items = len(items)
        n_flds  = sum(len(o.fields)    for o in items)
        n_ok    = sum(o.match_count    for o in items)
        n_bad   = sum(o.mismatch_count for o in items)
        has_mm  = any(o.has_mismatch   for o in items)
        status  = "⚠  MISMATCH" if has_mm else "✓  MATCHED"
        rfill   = _F_ROW_BAD if has_mm else _F_ROW_OK

        vals = [oid, n_items, n_flds, n_ok, n_bad, status]
        fts  = [_FT_NORMB, _FT_NORM, _FT_NORM, _FT_OK, _FT_BAD if has_mm else _FT_OK,
                _FT_BAD if has_mm else _FT_OK]
        als  = [_AL_C] * 6
        for col, (val, ft, al) in enumerate(zip(vals, fts, als), 1):
            _sc(ws.cell(row_idx, col, val), fill=rfill, font=ft, align=al)
        ws.row_dimensions[row_idx].height = 18

    for col in range(1, N + 1):
        ws.column_dimensions[get_column_letter(col)].width = [16, 14, 18, 16, 20, 16][col - 1]


def generate_comparison_report(
    results: List[OrderResult],
    path: str,
    start_time: datetime.datetime,
) -> None:
    """
    Write a full field-by-field comparison workbook.

    Sheet 1 – "Summary"   : executive KPI overview + per-order status table.
    Sheet 2 – "Full Detail": every order / every field, colour-coded match/mismatch.
    """
    ts       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    duration = int((datetime.datetime.now() - start_time).total_seconds())

    N     = 8
    HDRS  = ["Order ID", "Player / Line", "Exporter Tab", "Generator Tab",
             "Field", "Exporter Value", "Generator Value", "Status"]
    WITHS = [14, 26, 20, 20, 22, 28, 28, 14]

    wb     = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(ws_sum, results, ts, duration)

    ws = wb.create_sheet("Full Detail")
    ws.sheet_view.showGridLines = False
    _title_row(
        ws,
        "Falcon Stage – Order Exporter vs Generator: Full Comparison Detail",
        N, _F_HDR_CMP,
        (f"Order Range: {FALCON_ORDER_RANGE}  |  Channel: {FALCON_EXPORT_CHANNEL}  |  "
         f"Generated: {ts}  |  Duration: {duration}s"),
    )
    _stat_block(ws, [
        ("Total Line Items",  len(results),                             _F_STAT_TOT),
        ("Unique Order IDs",  len({r.order_id for r in results}),       _F_STAT_TOT),
        ("Fully Matched",     sum(1 for r in results if not r.has_mismatch), _F_STAT_OK),
        ("With Mismatches",   sum(1 for r in results if r.has_mismatch),     _F_STAT_BAD),
        ("Fields Matched",    sum(r.match_count    for r in results),        _F_STAT_OK),
        ("Fields Mismatched", sum(r.mismatch_count for r in results),        _F_STAT_BAD),
    ], start_row=4)

    _hdr_row(ws, HDRS, 7, _F_HDR_CMP)
    ws.freeze_panes = "A8"

    by_order: Dict[str, List[OrderResult]] = {}
    for r in results:
        by_order.setdefault(r.order_id, []).append(r)

    cur = 8
    for oid in sorted(by_order.keys(), key=_sort_key):
        order_items = by_order[oid]
        has_mm      = any(o.has_mismatch for o in order_items)
        banner_fill = _F_ORD_BAD if has_mm else _F_ORD_OK
        status_lbl  = "⚠  HAS MISMATCHES" if has_mm else "✓  ALL MATCHED"

        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=N)
        c = ws.cell(cur, 1, f"  Order #{oid}   —   {status_lbl}")
        c.fill = banner_fill; c.font = _FT_ORD; c.alignment = _AL_L
        ws.row_dimensions[cur].height = 22
        cur += 1

        for line_item in order_items:
            exp_tab, gen_tab, line_id = _parse_sheet_label(line_item.sheet)
            for fld in line_item.fields:
                mm       = not fld.matches
                rfill    = _F_ROW_BAD  if mm else _F_ROW_OK
                vfill    = _F_CELL_BAD if mm else _F_ROW_OK
                vfont    = _FT_BAD     if mm else _FT_NORM
                st_font  = _FT_BAD     if mm else _FT_OK
                st_lbl   = "MISMATCH"  if mm else "MATCH"

                row_data = [
                    (oid,                            _FT_NORMB,  rfill, _AL_C),
                    (line_id,                        _FT_SRC,    rfill, _AL_L),
                    (exp_tab,                        _FT_NORM,   rfill, _AL_C),
                    (gen_tab,                        _FT_NORM,   rfill, _AL_C),
                    (fld.column,                     _FT_NORM,   rfill, _AL_L),
                    (_safe_str(fld.exporter_value),  vfont,      vfill, _AL_L),
                    (_safe_str(fld.generator_value), vfont,      vfill, _AL_L),
                    (st_lbl,                         st_font,    rfill, _AL_C),
                ]
                for col, (val, ft, fl, al) in enumerate(row_data, 1):
                    _sc(ws.cell(cur, col, val), fill=fl, font=ft, align=al)
                ws.row_dimensions[cur].height = 16
                cur += 1

    for i, w in enumerate(WITHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.active = ws_sum
    wb.save(path)
    log.info("Comparison report saved → %s", path)


def generate_mismatches_report(
    results: List[OrderResult],
    path: str,
    start_time: datetime.datetime,
) -> None:
    """
    Write a focused mismatches-only workbook.

    Sheet 1 – "Summary"          : same executive overview as the comparison report.
    Sheet 2 – "Mismatches Detail": one row per mismatched field, self-described with
                                   a plain-English mismatch reason sentence.
    """
    ts         = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    duration   = int((datetime.datetime.now() - start_time).total_seconds())
    mismatched = [r for r in results if r.has_mismatch]
    total_bad  = sum(r.mismatch_count for r in mismatched)
    unique_oids = len({r.order_id for r in mismatched})

    N     = 8
    HDRS  = ["Order ID", "Player / Line", "Exporter Tab", "Generator Tab",
             "Mismatched Field", "Exporter Value", "Generator Value", "Mismatch Reason"]
    WITHS = [14, 26, 20, 20, 22, 28, 28, 54]

    wb     = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(ws_sum, results, ts, duration)

    ws = wb.create_sheet("Mismatches Detail")
    ws.sheet_view.showGridLines = False
    _title_row(
        ws,
        "Falcon Stage – Order Exporter vs Generator: Mismatches Report",
        N, _F_HDR_MM,
        (f"Order Range: {FALCON_ORDER_RANGE}  |  Channel: {FALCON_EXPORT_CHANNEL}  |  "
         f"Generated: {ts}  |  Duration: {duration}s"),
    )

    data_start = _stat_block(ws, [
        ("Orders with Mismatches",  unique_oids,              _F_STAT_BAD),
        ("Total Mismatched Fields", total_bad,                _F_STAT_BAD),
        ("Total Orders Checked",    len(results),             _F_STAT_TOT),
        ("Fully Matched Orders",    len(results) - len(mismatched), _F_STAT_OK),
    ], start_row=4)

    if not mismatched:
        msg_row = data_start + 1
        ws.merge_cells(start_row=msg_row, start_column=1, end_row=msg_row, end_column=N)
        c = ws.cell(msg_row, 1,
                    "✅  No mismatches found — every field matches between "
                    "Order Exporter and Order Generator.")
        c.fill = _F_STAT_OK
        c.font = Font(name="Calibri", bold=True, color="1A7A3C", size=12)
        c.alignment = _AL_C
        ws.row_dimensions[msg_row].height = 36
    else:
        hdr_row = data_start + 1
        _hdr_row(ws, HDRS, hdr_row, _F_HDR_MM)
        ws.freeze_panes = f"A{hdr_row + 1}"

        by_order: Dict[str, List[OrderResult]] = {}
        for r in mismatched:
            by_order.setdefault(r.order_id, []).append(r)

        cur = hdr_row + 1
        for oid in sorted(by_order.keys(), key=_sort_key):
            order_items = by_order[oid]
            total_mm    = sum(o.mismatch_count for o in order_items)

            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=N)
            c = ws.cell(cur, 1, f"  Order #{oid}   —   {total_mm} mismatched field(s)")
            c.fill = _F_ORD_BAD; c.font = _FT_ORD; c.alignment = _AL_L
            ws.row_dimensions[cur].height = 22
            cur += 1

            for line_item in order_items:
                bad = [f for f in line_item.fields if not f.matches]
                if not bad:
                    continue
                exp_tab, gen_tab, line_id = _parse_sheet_label(line_item.sheet)
                for fld in bad:
                    reason = (
                        f"Field '{fld.column}' differs — "
                        f"Order Generator recorded '{_safe_str(fld.generator_value)}'; "
                        f"Order Exporter recorded '{_safe_str(fld.exporter_value)}'"
                    )
                    row_data = [
                        (oid,                            _FT_NORMB, _F_ROW_BAD,  _AL_C),
                        (line_id,                        _FT_SRC,   _F_ROW_BAD,  _AL_L),
                        (exp_tab,                        _FT_NORM,  _F_ROW_BAD,  _AL_C),
                        (gen_tab,                        _FT_NORM,  _F_ROW_BAD,  _AL_C),
                        (fld.column,                     _FT_NORMB, _F_ROW_BAD,  _AL_L),
                        (_safe_str(fld.exporter_value),  _FT_BAD,   _F_CELL_BAD, _AL_L),
                        (_safe_str(fld.generator_value), _FT_BAD,   _F_CELL_BAD, _AL_L),
                        (reason,                         _FT_BAD,   _F_CELL_BAD, _AL_L),
                    ]
                    for col, (val, ft, fl, al) in enumerate(row_data, 1):
                        _sc(ws.cell(cur, col, val), fill=fl, font=ft, align=al)
                    ws.row_dimensions[cur].height = 16
                    cur += 1

    for i, w in enumerate(WITHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.active = ws_sum
    wb.save(path)
    log.info("Mismatches report saved → %s", path)


# ══════════════════════════════════════════════════════════════════════════════
# §6 – Browser automation (download both workbooks)
# ══════════════════════════════════════════════════════════════════════════════

def _trigger_download(page: Page, trigger_fn, save_path: Path) -> Path:
    """Execute *trigger_fn*, wait for the browser download, and persist the file."""
    with page.expect_download(timeout=60_000) as dl_info:
        trigger_fn()
    download: Download = dl_info.value
    download.save_as(str(save_path))
    log.info("Download saved → %s", save_path.name)
    return save_path


def _run_order_exporter(page: Page, ts: str) -> Path:
    """
    Navigate to the Order Exporter, configure it, trigger the Excel download,
    and return the saved file path.
    """
    log.info("Phase 1a: Launching Order Exporter …")
    page.goto(FALCON_APP_LAUNCHER_URL)
    page.wait_for_load_state("networkidle")

    with page.context.expect_page() as new_page_info:
        page.locator("div:nth-of-type(2) > a").first.click()
    export_page = new_page_info.value
    export_page.wait_for_load_state("networkidle")
    log.info("Order Exporter loaded: %s", export_page.url)

    export_page.locator("#exportType").select_option("orderNumbers")
    export_page.locator("#orderNumbers").fill(FALCON_ORDER_RANGE)
    export_page.locator("#channel").select_option(FALCON_EXPORT_CHANNEL)

    save_path = DOWNLOADS_DIR / f"order_exporter_{ts}.xlsx"
    _trigger_download(
        export_page,
        lambda: export_page.get_by_text("EXPORT TO EXCEL").click(),
        save_path,
    )
    export_page.close()
    return save_path


def _run_order_generator(page: Page, ts: str) -> Path:
    """
    Navigate to the Order Generator, configure it, trigger the Excel download,
    and return the saved file path.
    """
    log.info("Phase 1b: Launching Order Generator …")
    page.goto(FALCON_APP_LAUNCHER_URL)
    page.wait_for_load_state("networkidle")

    with page.context.expect_page() as new_page_info:
        page.locator("div:nth-of-type(1) > a").first.click()
    gen_page = new_page_info.value
    gen_page.wait_for_load_state("networkidle")
    log.info("Order Generator loaded: %s", gen_page.url)

    gen_page.locator("#orderName").fill(FALCON_PRODUCTION_ORDER)
    gen_page.locator("#orderNumbers").fill(FALCON_ORDER_RANGE)

    save_path = DOWNLOADS_DIR / f"order_generator_{ts}.xlsx"
    _trigger_download(
        gen_page,
        lambda: gen_page.locator("form > button").click(),
        save_path,
    )
    gen_page.close()
    return save_path


# ══════════════════════════════════════════════════════════════════════════════
# §7 – Module orchestrator
# ══════════════════════════════════════════════════════════════════════════════

def run() -> None:
    """
    Execute the full Falcon comparison pipeline.

    Steps
    -----
    1.  Launch browser and download Order Exporter workbook
    2.  Download Order Generator workbook
    3.  Close browser
    4.  Parse both workbooks
    5.  Compare field-by-field across all eligible tabs
    6.  Write Comparison Report XLSX
    7.  Write Mismatches Report XLSX
    8.  Write execution report (XlsxReporter)
    """
    reporter   = XlsxReporter(MODULE_NAME)
    start_time = datetime.datetime.now()
    ts         = start_time.strftime("%Y%m%d_%H%M%S")

    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

    exporter_path:  Optional[Path] = None
    generator_path: Optional[Path] = None

    reporter.add_step(
        "Module started",
        "INFO",
        (f"Order Range: {FALCON_ORDER_RANGE}  |  "
         f"Channel: {FALCON_EXPORT_CHANNEL}  |  "
         f"Launcher: {FALCON_APP_LAUNCHER_URL}"),
    )

    # ── Phase 1: Browser automation ───────────────────────────────────────────
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=400)
        context = browser.new_context(viewport=VIEWPORT, accept_downloads=True)
        page    = context.new_page()

        try:
            exporter_path = _run_order_exporter(page, ts)
            reporter.add_step(
                "Download Order Exporter workbook",
                "PASS",
                exporter_path.name,
            )

            generator_path = _run_order_generator(page, ts)
            reporter.add_step(
                "Download Order Generator workbook",
                "PASS",
                generator_path.name,
            )
        except Exception as exc:
            reporter.add_step("Browser automation failed", "FAIL", str(exc))
            reporter.save()
            raise
        finally:
            context.close()
            browser.close()
            log.info("Browser closed.")

    reporter.add_step("Browser session closed", "PASS")

    # ── Phase 2: Parse workbooks ──────────────────────────────────────────────
    try:
        exporter_sheets  = parse_excel_workbook(exporter_path)
        generator_sheets = parse_excel_workbook(generator_path)
        reporter.add_step(
            "Parse both Excel workbooks",
            "PASS",
            (f"Exporter tabs: {len(exporter_sheets)}  |  "
             f"Generator tabs: {len(generator_sheets)}"),
        )
    except Exception as exc:
        reporter.add_step("Workbook parsing failed", "FAIL", str(exc))
        reporter.save()
        raise

    # ── Phase 3: Compare ──────────────────────────────────────────────────────
    try:
        results = compare_workbooks(exporter_sheets, generator_sheets)
        mismatched_count = sum(1 for r in results if r.has_mismatch)
        matched_count    = len(results) - mismatched_count
        reporter.add_step(
            "Compare workbooks field-by-field",
            "PASS" if mismatched_count == 0 else "FAIL",
            (f"Line items: {len(results)}  |  "
             f"Matched: {matched_count}  |  "
             f"Mismatched: {mismatched_count}"),
        )
    except Exception as exc:
        reporter.add_step("Comparison engine failed", "FAIL", str(exc))
        reporter.save()
        raise

    # ── Phase 4: Write reports ────────────────────────────────────────────────
    cmp_path = str(REPORTS_DIR / f"Falcon_Stage_Order_Comparison_Report_{ts}.xlsx")
    mm_path  = str(REPORTS_DIR / f"Falcon_Stage_Mismatches_Report_{ts}.xlsx")

    generate_comparison_report(results, cmp_path, start_time)
    reporter.add_step("Write Comparison Report", "PASS", cmp_path)

    generate_mismatches_report(results, mm_path, start_time)
    reporter.add_step("Write Mismatches Report", "PASS", mm_path)

    status = "PASS" if mismatched_count == 0 else "FAIL"
    reporter.add_step("Module completed", status,
                      f"Matched: {matched_count}  |  Mismatched: {mismatched_count}")

    report_path = reporter.save()
    log.info("Execution report → %s", report_path)
    log.info("Comparison report → %s", cmp_path)
    log.info("Mismatches report → %s", mm_path)


if __name__ == "__main__":
    run()
