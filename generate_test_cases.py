"""
Generate professional test cases XLSX for Flite Sports Bundle Product Flow.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# ──────────────────────────────────────────────────────────
# DATA
# ──────────────────────────────────────────────────────────

MODULES = [
    {
        "name": "M01 – Authentication",
        "tab_color": "1F4E79",
        "cases": [
            ("TC-AUTH-01", "Successful login with valid credentials", "App is accessible", "Enter valid email & password → click Login → click \"Let's Go\"", "Dashboard loads successfully", "Happy"),
            ("TC-AUTH-02", "Login with invalid email format", "—", "Enter \"adminbitcot.com\" (no @) → submit", "Validation error: 'Enter valid email'", "Negative"),
            ("TC-AUTH-03", "Login with wrong password", "Valid email exists", "Enter correct email, wrong password → submit", "Error: 'Invalid credentials' / login fails", "Negative"),
            ("TC-AUTH-04", "Login with empty email field", "—", "Leave email blank, enter password → submit", "Validation error on email field", "Negative"),
            ("TC-AUTH-05", "Login with empty password field", "—", "Enter email, leave password blank → submit", "Validation error on password field", "Negative"),
            ("TC-AUTH-06", "Login with both fields empty", "—", "Click submit without any input", "Both fields show validation error", "Negative"),
            ("TC-AUTH-07", "Login with SQL injection in email", "—", "Enter \"' OR 1=1 --\" in email field → submit", "Login fails; no DB error exposed", "Security"),
            ("TC-AUTH-08", "Login with XSS payload in password", "—", "Enter <script>alert(1)</script> as password", "Input sanitized; no script executes", "Security"),
            ("TC-AUTH-09", "Password field masked by default", "—", "Observe password field on load", "Characters are masked (type=password)", "Happy"),
            ("TC-AUTH-10", "Session persistence after login", "Successfully logged in", "Close and reopen browser tab", "Session remains active (or re-auth required per policy)", "Positive"),
            ("TC-AUTH-11", "Login with deactivated admin account", "Account disabled in backend", "Enter valid credentials of disabled user", "Error: 'Account disabled' or access denied", "Negative"),
            ("TC-AUTH-12", "CapsLock warning on password field", "—", "Enable CapsLock, focus password field", "CapsLock warning indicator displayed", "Positive"),
        ]
    },
    {
        "name": "M02 – Partners & Clubs",
        "tab_color": "1A5276",
        "cases": [
            ("TC-PART-01", "Navigate to PARTNERS/CLUBS from sidebar", "Logged in", "Click PARTNERS/CLUBS in nav", "Partners list page loads", "Happy"),
            ("TC-PART-02", "Search for existing partner by name", "Partners page open", "Type 'QA partn' in search box", "Matching partners shown in table", "Happy"),
            ("TC-PART-03", "Search for existing partner 'karen'", "Partners page open", "Type 'karen' in search box", "Karen's record appears in results", "Happy"),
            ("TC-PART-04", "Search with non-existent partner name", "Partners page open", "Type 'xyznonexistent999'", "Empty results / 'No records found'", "Negative"),
            ("TC-PART-05", "Search with special characters", "Partners page open", "Type @#$%^&* in search", "No crash; empty or sanitized result", "Negative"),
            ("TC-PART-06", "Search with empty string after clearing", "Partners page open", "Type then clear search box", "Full partner list restored", "Positive"),
            ("TC-PART-07", "Search with partial name (case-insensitive)", "Partners page open", "Type 'KAREN' (uppercase)", "Karen's record still appears", "Positive"),
            ("TC-PART-08", "Open partner detail via manage/edit button", "Partner found in search", "Click manage button (column 11) on partner row", "Partner detail/management page opens", "Happy"),
            ("TC-PART-09", "Navigate to sub-tabs inside partner detail", "Partner detail open", "Click BUNDLES, PRODUCTS tabs", "Each respective tab loads correct content", "Happy"),
            ("TC-PART-10", "Navigate away from partner without saving", "Partner detail open, unsaved changes", "Click PARTNERS/CLUBS in sidebar", "Warning prompt shown OR changes lost (verify behavior)", "Edge"),
            ("TC-PART-11", "Pagination on partners list", ">10 partners exist", "Scroll/navigate to next page", "Next set of partners loaded correctly", "Positive"),
            ("TC-PART-12", "Sort partners by column header", "Partners list loaded", "Click on column header (e.g., Name)", "Table sorted ascending/descending", "Positive"),
        ]
    },
    {
        "name": "M03 – Bundle Listing & View",
        "tab_color": "154360",
        "cases": [
            ("TC-BUND-01", "View BUNDLES tab for a partner with existing bundles", "Partner with bundles", "Open partner → click BUNDLES", "Bundle grid displays all bundles", "Happy"),
            ("TC-BUND-02", "View BUNDLES tab for a partner with no bundles", "Partner has no bundles", "Open partner → click BUNDLES", "Empty state message shown (e.g., 'No bundles yet')", "Positive"),
            ("TC-BUND-03", "View bundle details via View button", "Bundles listed", "Click view (eye) icon on a bundle card", "Bundle detail view opens with correct data", "Happy"),
            ("TC-BUND-04", "View button multiple rapid clicks", "Bundles listed", "Click view icon 7–8 times rapidly", "No duplicate dialogs; single view opens", "Edge"),
            ("TC-BUND-05", "Bundle card displays correct product count", "Bundle exists", "View bundle grid", "Product names and count match created bundle", "Positive"),
            ("TC-BUND-06", "Bundle card displays correct image", "Bundle with image exists", "View bundle grid", "Thumbnail image renders without broken link", "Positive"),
            ("TC-BUND-07", "Bundle card shows correct price", "Bundle with price set", "View bundle grid", "Price displayed matches saved value ($78.99)", "Positive"),
        ]
    },
    {
        "name": "M04 – Bundle Creation (Happy)",
        "tab_color": "145A32",
        "cases": [
            ("TC-BC-01", "Create bundle with all required fields", "Partner open, on BUNDLES tab", "Click CREATE BUNDLE → select products → set size chart title → upload size chart images → click NEXT → upload thumbnail → click SAVE BUNDLE", "Bundle created and appears in bundle grid", "Happy"),
            ("TC-BC-02", "Create bundle with LOCAL CUSTOMIZATION product", "Bundle modal open", "Search and select LOCAL CUSTOMIZATION product → select variants → proceed", "Variants listed in bundle correctly", "Happy"),
            ("TC-BC-03", "Create bundle with OFF THE SHELF product", "Bundle modal open", "Search and add OFF THE SHELF product → select variants", "Both product types coexist in bundle", "Happy"),
            ("TC-BC-04", "Create bundle with valid price ($78.99)", "Bundle modal Step 2", "Enter 78.99 in BUNDLE PRICE field → save", "Bundle saved with correct price shown", "Happy"),
            ("TC-BC-05", "Set SIZE TITLE for bundle", "Bundle modal open, products selected", "Enter 'SIZE CHARt' in SIZE TITLE field", "Title saved and displayed in size chart section", "Happy"),
            ("TC-BC-06", "Upload valid size chart image (PNG)", "Bundle creation modal", "Upload image (PNG < size limit) in size chart section", "Image preview shown; no error", "Happy"),
            ("TC-BC-07", "Upload valid bundle thumbnail image", "Bundle Step 2 (details page)", "Click upload area → select valid PNG", "Thumbnail preview displayed; no error", "Happy"),
            ("TC-BC-08", "Add multiple size chart images", "Bundle modal, size chart section", "Upload first image → click '+ ADD SIZE CHART' → upload second", "Both images displayed in size chart list", "Happy"),
            ("TC-BC-09", "NEXT button transitions to Step 2", "Step 1 complete with products selected", "Click NEXT", "Step 2 (details) page opens correctly", "Happy"),
            ("TC-BC-10", "Create bundle with bundle price = 0", "Bundle modal Step 2", "Enter 0.00 in BUNDLE PRICE → save", "Bundle saved (or validation error if 0 not allowed)", "Edge"),
        ]
    },
    {
        "name": "M05 – Bundle Creation (Negative)",
        "tab_color": "7B241C",
        "cases": [
            ("TC-BCN-01", "Save bundle without selecting any product", "Bundle modal open", "Click NEXT without selecting products", "Validation error: 'Select at least one product'", "Negative"),
            ("TC-BCN-02", "Save bundle without uploading thumbnail image", "Bundle Step 2 open", "Click SAVE BUNDLE without uploading image", "Validation error: 'Bundle image required'", "Negative"),
            ("TC-BCN-03", "Save bundle without setting bundle price", "Bundle Step 2 open", "Leave BUNDLE PRICE empty → click SAVE BUNDLE", "Validation error: 'Price is required'", "Negative"),
            ("TC-BCN-04", "Enter negative price in BUNDLE PRICE", "Bundle Step 2 open", "Type '-50' in price field → save", "Validation error: 'Price must be positive'", "Negative"),
            ("TC-BCN-05", "Enter alphabetic value in BUNDLE PRICE", "Bundle Step 2 open", "Type 'abc' in price field → save", "Field rejects non-numeric input OR validation error", "Negative"),
            ("TC-BCN-06", "Enter extremely large price value", "Bundle Step 2 open", "Type '99999999999' → save", "Validation error or max value enforced", "Negative"),
            ("TC-BCN-07", "Upload non-image file as size chart", "Size chart upload area", "Attempt to upload a .pdf or .exe file", "Error: 'Only image files allowed'", "Negative"),
            ("TC-BCN-08", "Upload oversized image as bundle thumbnail", "Bundle Step 2", "Upload image exceeding size limit (e.g., 20MB)", "Error: 'File size exceeds limit'", "Negative"),
            ("TC-BCN-09", "Create bundle without setting SIZE TITLE", "Bundle modal, products selected", "Leave SIZE TITLE blank → click NEXT", "Validation error if title is required", "Negative"),
            ("TC-BCN-10", "Create bundle with duplicate name as existing bundle", "Bundle modal", "Enter same title as an existing bundle → save", "Error: 'Bundle with this name already exists'", "Negative"),
            ("TC-BCN-11", "Close bundle modal mid-creation without saving", "Bundle modal partially filled", "Click Close (X) button", "Confirmation prompt: 'Unsaved changes will be lost'", "Negative"),
            ("TC-BCN-12", "Click NEXT on Step 1 with no variant selected", "Products shown but none checked", "Click NEXT button", "Error: 'Please select at least one variant'", "Negative"),
            ("TC-BCN-13", "Upload corrupt/invalid image file", "Bundle thumbnail upload", "Upload a corrupted .png file", "Error: 'Invalid file / could not process image'", "Negative"),
            ("TC-BCN-14", "Click SAVE BUNDLE multiple times rapidly", "Bundle Step 2, all fields filled", "Double/triple click SAVE BUNDLE", "Bundle created only once; no duplicates", "Edge"),
            ("TC-BCN-15", "Network failure during bundle save", "Bundle modal filled", "Disconnect network → click SAVE BUNDLE", "Error message shown; bundle NOT partially created", "Negative"),
            ("TC-BCN-16", "Create bundle for partner with no products assigned", "Partner with no products", "Open CREATE BUNDLE", "Empty product list with 'No products available' message", "Negative"),
            ("TC-BCN-17", "Search for non-existent product in bundle modal", "Bundle modal open", "Type 'xyz_nonexistent' in product search", "'No products found' message shown", "Negative"),
            ("TC-BCN-18", "Add same product to bundle twice", "Bundle modal open", "Select same product twice from search", "Duplicate prevented OR warning shown", "Negative"),
            ("TC-BCN-19", "Price field with special characters ($, comma)", "Bundle Step 2", "Type '$78,99' in price field → save", "Input sanitized to valid decimal OR validation error", "Negative"),
            ("TC-BCN-20", "Remove size chart image then save without adding new one", "Size chart image uploaded, then removed", "Click Remove → click NEXT → save", "Validation if size chart image is mandatory", "Negative"),
        ]
    },
    {
        "name": "M06 – Bundle Editing",
        "tab_color": "784212",
        "cases": [
            ("TC-BE-01", "Open edit modal for existing bundle", "Bundle exists in grid", "Click Edit (pencil) button on bundle card", "Bundle edit modal opens with pre-filled data", "Happy"),
            ("TC-BE-02", "Edit bundle price and save", "Bundle edit modal open", "Change price value → click SAVE BUNDLE", "Price updated correctly in bundle grid", "Happy"),
            ("TC-BE-03", "Edit SIZE TITLE and save", "Bundle edit modal open", "Change SIZE TITLE → save", "Updated title reflected in bundle", "Happy"),
            ("TC-BE-04", "Replace bundle thumbnail image", "Bundle edit modal Step 2", "Upload new thumbnail → save", "New image displayed on bundle card", "Happy"),
            ("TC-BE-05", "Add additional product to existing bundle", "Bundle edit modal open", "Search and add new product → save", "Bundle now contains additional product", "Happy"),
            ("TC-BE-06", "Remove a product variant from bundle during edit", "Bundle edit modal, variant selected", "Deselect a variant → save", "Variant removed from bundle", "Happy"),
            ("TC-BE-07", "Close edit modal without saving changes", "Bundle edit modal with changes", "Click Close (X)", "Changes discarded; bundle unchanged", "Positive"),
            ("TC-BE-08", "Edit bundle price to empty and save", "Bundle edit modal", "Clear price field → click SAVE BUNDLE", "Validation error: price required", "Negative"),
            ("TC-BE-09", "Edit bundle price to negative value", "Bundle edit modal", "Enter -1 → save", "Validation error", "Negative"),
            ("TC-BE-10", "Edit bundle and upload invalid image format", "Bundle edit modal Step 2", "Upload .gif or .bmp (if unsupported) → save", "Error shown; previous image retained", "Negative"),
            ("TC-BE-11", "Edit bundle with all products deselected", "Bundle edit modal", "Uncheck all products → click NEXT", "Validation error: must have at least one product", "Negative"),
            ("TC-BE-12", "Edit and save without making any changes", "Bundle edit modal open", "Open, change nothing → click SAVE BUNDLE", "Bundle saved without error; data unchanged", "Edge"),
            ("TC-BE-13", "Concurrent edit of same bundle by two users", "Two sessions open", "Both users open edit for same bundle → both save", "Last-write-wins or conflict error shown", "Edge"),
        ]
    },
    {
        "name": "M07 – Bundle Deletion",
        "tab_color": "641E16",
        "cases": [
            ("TC-BD-01", "Delete a bundle with confirmation", "Bundle exists", "Click Delete (trash) button → confirm dialog → click Confirm/Danger button", "Bundle removed from grid", "Happy"),
            ("TC-BD-02", "Cancel bundle deletion", "Bundle exists", "Click Delete → on confirmation dialog click Cancel", "Bundle NOT deleted; remains in grid", "Positive"),
            ("TC-BD-03", "Delete last remaining bundle for a partner", "Only one bundle exists", "Delete it → confirm", "Bundle deleted; empty state shown", "Positive"),
            ("TC-BD-04", "Delete bundle already synced to Shopify", "Synced bundle exists", "Click Delete → confirm", "Warning: 'Bundle is live on Shopify — confirm deletion' OR bundle deleted with Shopify update", "Edge"),
            ("TC-BD-05", "Delete multiple bundles sequentially", "Multiple bundles exist", "Delete first → confirm → delete second → confirm", "Both deleted successfully", "Happy"),
            ("TC-BD-06", "Rapid multiple clicks on Delete button", "Bundle exists", "Click Delete icon multiple times quickly", "Confirmation appears only once; no duplicate modals", "Edge"),
            ("TC-BD-07", "Delete bundle during active order", "Bundle tied to an active Shopify order", "Click Delete", "Error: 'Bundle is part of an active order; cannot delete'", "Negative"),
            ("TC-BD-08", "Browser back after deleting bundle", "Bundle deleted", "Press browser back button", "Previous state does not restore deleted bundle", "Edge"),
        ]
    },
    {
        "name": "M08 – Bundle Sync to Shopify",
        "tab_color": "0B5345",
        "cases": [
            ("TC-SYNC-01", "Sync bundle to Shopify successfully", "Bundle saved, valid Shopify connection", "Click SYNC BUNDLE", "Success message; bundle live on Shopify", "Happy"),
            ("TC-SYNC-02", "Sync bundle without thumbnail image", "Bundle missing image", "Click SYNC BUNDLE", "Error: 'Image required before sync'", "Negative"),
            ("TC-SYNC-03", "Sync bundle without price set", "Bundle missing price", "Click SYNC BUNDLE", "Error: 'Price required before sync'", "Negative"),
            ("TC-SYNC-04", "Sync bundle with Shopify disconnected/API error", "Shopify connection broken", "Click SYNC BUNDLE", "Error: 'Failed to sync — check Shopify connection'", "Negative"),
            ("TC-SYNC-05", "Sync same bundle twice", "Bundle already synced", "Click SYNC BUNDLE again", "Success: bundle updated on Shopify OR 'No changes to sync'", "Positive"),
            ("TC-SYNC-06", "Sync bundle with invalid product SKU", "Product has invalid/missing SKU", "Click SYNC BUNDLE", "Error identifying the problematic product", "Negative"),
            ("TC-SYNC-07", "Sync products (SYNC PRODUCTS button)", "On PRODUCTS tab", "Click SYNC PRODUCTS → confirm", "Products synced; success notification shown", "Happy"),
            ("TC-SYNC-08", "Cancel sync confirmation dialog", "SYNC BUNDLE clicked", "Click Cancel on confirmation", "Sync cancelled; no changes to Shopify", "Positive"),
        ]
    },
    {
        "name": "M09 – Product Management",
        "tab_color": "1A5276",
        "cases": [
            ("TC-PROD-01", "View LOCAL CUSTOMIZATION products tab", "Partner detail, PRODUCTS tab", "Click LOCAL CUSTOMIZATION filter", "Only local customization products shown", "Happy"),
            ("TC-PROD-02", "View SUBLIMATION products tab", "Partner detail, PRODUCTS tab", "Click SUBLIMATION tab", "Only sublimation products shown", "Happy"),
            ("TC-PROD-03", "Sync products successfully", "On partner PRODUCTS tab", "Click SYNC PRODUCTS → confirm dialog", "Products synced; list refreshed", "Happy"),
            ("TC-PROD-04", "Products tab shows empty state if no products assigned", "Partner has no products", "Navigate to PRODUCTS tab", "'No products' / empty state message shown", "Positive"),
            ("TC-PROD-05", "Search for product within partner products", "Partner has multiple products", "Use search/filter", "Matching products displayed", "Positive"),
        ]
    },
    {
        "name": "M10 – Size Chart Management",
        "tab_color": "4A235A",
        "cases": [
            ("TC-SC-01", "Add size chart with valid title and image", "Bundle modal, products selected", "Enter SIZE TITLE → upload image → click '+ ADD SIZE CHART'", "Size chart entry added to list", "Happy"),
            ("TC-SC-02", "Add multiple size charts to one bundle", "Bundle modal", "Add two separate size chart images", "Both displayed in size chart section", "Happy"),
            ("TC-SC-03", "Remove an uploaded size chart image", "Size chart image uploaded", "Click Remove button on uploaded image", "Image removed; slot reset to upload state", "Happy"),
            ("TC-SC-04", "Add size chart without entering a title", "Bundle modal", "Leave SIZE TITLE empty → upload image", "Validation error if title is required", "Negative"),
            ("TC-SC-05", "Add size chart without uploading image", "Bundle modal", "Enter title → click '+ ADD SIZE CHART' without uploading", "Validation error: 'Image required for size chart'", "Negative"),
            ("TC-SC-06", "Upload non-PNG file as size chart", "Size chart upload", "Upload a .doc or .pdf file", "Error: 'Only PNG/JPG allowed'", "Negative"),
            ("TC-SC-07", "Enter extremely long SIZE TITLE", "Bundle modal", "Enter 500+ character title", "Input truncated or validation error shown", "Negative"),
            ("TC-SC-08", "Enter special characters in SIZE TITLE", "Bundle modal", "Enter <script> or '; DROP TABLE as title", "Input sanitized; no security issue", "Security"),
        ]
    },
    {
        "name": "M11 – UI & UX Edge Cases",
        "tab_color": "2C3E50",
        "cases": [
            ("TC-UI-01", "Sidebar navigation highlights active section", "Any page", "Click BUNDLES tab", "BUNDLES nav item highlighted/active", "Positive"),
            ("TC-UI-02", "Browser back/forward during bundle creation", "Bundle modal open", "Press browser back", "Confirm leave dialog OR modal closes without crashing", "Edge"),
            ("TC-UI-03", "Page refresh during bundle creation (Step 1)", "Bundle modal partially filled", "Press F5", "Modal closes; unsaved data lost (expected behavior)", "Edge"),
            ("TC-UI-04", "Viewport responsiveness at 1520×695", "Any page", "Open at defined viewport", "All elements visible; no overflow", "Positive"),
            ("TC-UI-05", "Tab key navigation through bundle form", "Bundle modal open", "Press Tab through all fields", "Focus moves in logical order", "Accessibility"),
            ("TC-UI-06", "Modal overlay closes on X button", "Any modal open", "Click X / Close button", "Modal dismisses cleanly", "Happy"),
            ("TC-UI-07", "Double-click on bundle create button", "BUNDLES tab", "Double-click CREATE BUNDLE", "Only one modal opens", "Edge"),
            ("TC-UI-08", "Search box retains value after navigation", "Partner search done", "Navigate away and return", "Search cleared OR retained (verify design intent)", "Edge"),
            ("TC-UI-09", "Table row click selects correct partner", "Multiple partners in search results", "Click row 13", "Correct partner (row 13) detail opens", "Positive"),
            ("TC-UI-10", "Verify loading indicators during async operations", "Slow network", "Trigger sync or save", "Loading spinner/indicator shown during operation", "Positive"),
        ]
    },
    {
        "name": "M12 – Cross-Partner Isolation",
        "tab_color": "117A65",
        "cases": [
            ("TC-ISO-01", "Bundles of Partner A not visible under Partner B", "Two partners with bundles", "Switch between partners' BUNDLES tabs", "Each partner sees only their own bundles", "Happy"),
            ("TC-ISO-02", "Create bundle under Partner A does not affect Partner B", "Bundle created for Partner A", "Check Partner B's BUNDLES", "Partner B unaffected", "Happy"),
            ("TC-ISO-03", "Delete bundle under Partner A does not delete Partner B's", "Delete from Partner A", "Check Partner B's BUNDLES", "Partner B's bundles intact", "Happy"),
            ("TC-ISO-04", "Direct URL access to another partner's bundle", "Bundle URL known", "Manually navigate to another partner's bundle URL", "Access denied OR redirect to authorized area", "Security"),
        ]
    },
]

# ──────────────────────────────────────────────────────────
# STYLE HELPERS
# ──────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

TYPE_STYLE = {
    "Happy":        {"fill": "D5F5E3", "font_color": "1E8449", "label": "✔ Happy Path"},
    "Positive":     {"fill": "D6EAF8", "font_color": "1A5276", "label": "✔ Positive"},
    "Negative":     {"fill": "FADBD8", "font_color": "922B21", "label": "✘ Negative"},
    "Edge":         {"fill": "FEF9E7", "font_color": "9A7D0A", "label": "⚠ Edge Case"},
    "Security":     {"fill": "F5EEF8", "font_color": "6C3483", "label": "🔒 Security"},
    "Accessibility":{"fill": "EBF5FB", "font_color": "1B4F72", "label": "♿ Accessibility"},
}

HEADER_COLS = ["TC ID", "Test Case Title", "Precondition", "Test Steps", "Expected Result", "Type"]
COL_WIDTHS  = [14, 48, 32, 58, 48, 16]

# ──────────────────────────────────────────────────────────
# COVER SHEET
# ──────────────────────────────────────────────────────────

def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.sheet_view.showGridLines = False

    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Background
    bg = hex_fill("1F4E79")
    for row in range(1, 40):
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = bg

    # Title block
    ws.merge_cells("B3:H4")
    c = ws["B3"]
    c.value = "FLITE SPORTS – BUNDLE PRODUCT FLOW"
    c.font = Font(name="Calibri", bold=True, size=22, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:H5")
    c = ws["B5"]
    c.value = "Comprehensive Test Case Document"
    c.font = Font(name="Calibri", size=14, color="AED6F1")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B6:H6")
    c = ws["B6"]
    c.value = "Shopify / E-Commerce Admin Platform"
    c.font = Font(name="Calibri", size=11, color="AED6F1", italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Divider row
    ws.row_dimensions[8].height = 4
    for col in range(2, 9):
        ws.cell(row=8, column=col).fill = hex_fill("AED6F1")

    # Stats
    total = sum(len(m["cases"]) for m in MODULES)
    type_counts = {}
    for m in MODULES:
        for c in m["cases"]:
            t = c[5]
            type_counts[t] = type_counts.get(t, 0) + 1

    info_rows = [
        ("Project",        "Flite Sports – Stage Environment"),
        ("Application URL","https://flitesports-stage.bitcotapps.com"),
        ("Test Scope",     "Bundle Product Flow (Admin Panel)"),
        ("Total Test Cases", str(total)),
        ("Total Modules",  str(len(MODULES))),
        ("Date",           "2026-06-11"),
        ("Prepared By",    "QA Team"),
        ("Version",        "1.0"),
    ]

    row_start = 10
    for i, (label, value) in enumerate(info_rows):
        r = row_start + i
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"B{r}:C{r}")
        ws.merge_cells(f"D{r}:H{r}")
        lbl = ws[f"B{r}"]
        lbl.value = label
        lbl.font = Font(name="Calibri", bold=True, size=11, color="AED6F1")
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val = ws[f"D{r}"]
        val.value = value
        val.font = Font(name="Calibri", size=11, color="FFFFFF")
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Coverage breakdown
    ws.row_dimensions[row_start + len(info_rows) + 1].height = 4
    for col in range(2, 9):
        ws.cell(row=row_start + len(info_rows) + 1, column=col).fill = hex_fill("AED6F1")

    r2 = row_start + len(info_rows) + 3
    ws.merge_cells(f"B{r2}:H{r2}")
    h = ws[f"B{r2}"]
    h.value = "Test Type Coverage Summary"
    h.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r2].height = 26

    r2 += 1
    headers = ["Type", "Count", "", "Type", "Count"]
    col_positions = [2, 3, 4, 5, 6]
    for cp, hv in zip(col_positions, headers):
        c2 = ws.cell(row=r2, column=cp)
        c2.value = hv
        c2.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        c2.fill = hex_fill("AED6F1")
        c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r2].height = 20

    types_list = list(type_counts.items())
    half = (len(types_list) + 1) // 2
    for i in range(half):
        r3 = r2 + 1 + i
        ws.row_dimensions[r3].height = 20
        # Left column
        t1, cnt1 = types_list[i]
        style1 = TYPE_STYLE.get(t1, {"fill": "FFFFFF", "font_color": "000000"})
        lc1 = ws.cell(row=r3, column=2)
        lc1.value = TYPE_STYLE.get(t1, {}).get("label", t1)
        lc1.font = Font(name="Calibri", size=10, color=style1["font_color"], bold=True)
        lc1.fill = hex_fill(style1["fill"])
        lc1.alignment = Alignment(horizontal="center", vertical="center")
        lc2 = ws.cell(row=r3, column=3)
        lc2.value = cnt1
        lc2.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
        lc2.alignment = Alignment(horizontal="center", vertical="center")
        # Right column
        if i + half < len(types_list):
            t2, cnt2 = types_list[i + half]
            style2 = TYPE_STYLE.get(t2, {"fill": "FFFFFF", "font_color": "000000"})
            rc1 = ws.cell(row=r3, column=5)
            rc1.value = TYPE_STYLE.get(t2, {}).get("label", t2)
            rc1.font = Font(name="Calibri", size=10, color=style2["font_color"], bold=True)
            rc1.fill = hex_fill(style2["fill"])
            rc1.alignment = Alignment(horizontal="center", vertical="center")
            rc2 = ws.cell(row=r3, column=6)
            rc2.value = cnt2
            rc2.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
            rc2.alignment = Alignment(horizontal="center", vertical="center")

    # Module index
    r4 = r2 + half + 3
    ws.merge_cells(f"B{r4}:H{r4}")
    mi = ws[f"B{r4}"]
    mi.value = "Module Index"
    mi.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    mi.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r4].height = 26

    r4 += 1
    for i, mod in enumerate(MODULES):
        r5 = r4 + i
        ws.row_dimensions[r5].height = 19
        ws.merge_cells(f"B{r5}:C{r5}")
        ws.merge_cells(f"D{r5}:F{r5}")
        ws.merge_cells(f"G{r5}:H{r5}")
        n = ws.cell(row=r5, column=2)
        n.value = mod["name"]
        n.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        n.fill = hex_fill(mod["tab_color"])
        n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cnt = ws.cell(row=r5, column=4)
        cnt.value = f"{len(mod['cases'])} test cases"
        cnt.font = Font(name="Calibri", size=10, color="D5D8DC")
        cnt.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ──────────────────────────────────────────────────────────
# MODULE SHEETS
# ──────────────────────────────────────────────────────────

def build_module_sheet(wb, module):
    ws = wb.create_sheet(title=module["name"][:31])
    ws.sheet_properties.tabColor = module["tab_color"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title row ──
    ws.merge_cells(f"A1:{get_column_letter(len(HEADER_COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = module["name"].upper()
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = hex_fill(module["tab_color"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 32

    # ── Sub-header ──
    ws.merge_cells(f"A2:{get_column_letter(len(HEADER_COLS))}2")
    sub = ws["A2"]
    total = len(module["cases"])
    neg   = sum(1 for c in module["cases"] if c[5] == "Negative")
    happy = sum(1 for c in module["cases"] if c[5] == "Happy")
    sub.value = f"  Total: {total} Test Cases   |   Happy Path: {happy}   |   Negative: {neg}"
    sub.font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    sub.fill = hex_fill("2C3E50")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 18

    # ── Column headers ──
    header_fill = hex_fill("1F3864")
    for col_idx, col_name in enumerate(HEADER_COLS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_name
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[3].height = 26

    # ── Data rows ──
    for row_idx, case in enumerate(module["cases"], start=4):
        tc_id, title, precond, steps, expected, tc_type = case
        style = TYPE_STYLE.get(tc_type, {"fill": "FFFFFF", "font_color": "000000"})

        row_fill_hex = "F8F9FA" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = hex_fill(row_fill_hex)

        values = [tc_id, title, precond, steps, expected, tc_type]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

            # Type badge column
            if col_idx == 6:
                cell.value = TYPE_STYLE.get(tc_type, {}).get("label", tc_type)
                cell.fill = hex_fill(style["fill"])
                cell.font = Font(name="Calibri", size=9, bold=True, color=style["font_color"])
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_idx == 1:
                cell.fill = hex_fill("EBF5FB")
                cell.font = Font(name="Calibri", size=10, bold=True, color="1A5276")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=10, color="2C3E50")

        ws.row_dimensions[row_idx].height = 52

    # ── Footer ──
    footer_row = len(module["cases"]) + 4
    ws.merge_cells(f"A{footer_row}:{get_column_letter(len(HEADER_COLS))}{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = f"  End of {module['name']}  |  {len(module['cases'])} test cases documented"
    footer.font = Font(name="Calibri", size=9, color="FFFFFF", italic=True)
    footer.fill = hex_fill("2C3E50")
    footer.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[footer_row].height = 18


# ──────────────────────────────────────────────────────────
# SUMMARY SHEET
# ──────────────────────────────────────────────────────────

def build_summary(wb):
    ws = wb.create_sheet(title="Summary", index=1)
    ws.sheet_properties.tabColor = "2C3E50"
    ws.sheet_view.showGridLines = False

    col_widths = [6, 28, 14, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = "TEST CASES SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = hex_fill("1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Flite Sports Bundle Product Flow  |  All Modules"
    ws["A2"].font = Font(name="Calibri", size=11, color="AED6F1", italic=True)
    ws["A2"].fill = hex_fill("2C3E50")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Header
    headers = ["#", "Module", "Total TCs", "Happy", "Positive", "Negative", "Edge", "Security"]
    header_fill = hex_fill("1F3864")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.row_dimensions[4].height = 26

    totals = {"Happy": 0, "Positive": 0, "Negative": 0, "Edge": 0, "Security": 0, "Accessibility": 0}
    grand_total = 0

    for row_idx, module in enumerate(MODULES, start=5):
        counts = {"Happy": 0, "Positive": 0, "Negative": 0, "Edge": 0, "Security": 0, "Accessibility": 0}
        for case in module["cases"]:
            t = case[5]
            counts[t] = counts.get(t, 0) + 1
            totals[t]  = totals.get(t, 0)  + 1
        total = len(module["cases"])
        grand_total += total

        row_fill = hex_fill("F8F9FA" if row_idx % 2 == 0 else "FFFFFF")
        row_values = [row_idx - 4, module["name"], total,
                      counts["Happy"], counts["Positive"],
                      counts["Negative"], counts["Edge"],
                      counts["Security"]]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 22
            if col_idx == 1:
                cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
                cell.fill = hex_fill(module["tab_color"])
            elif col_idx == 2:
                cell.font = Font(name="Calibri", size=10, bold=True, color="2C3E50")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif col_idx == 3:
                cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
                cell.fill = row_fill
            elif col_idx == 6 and val > 0:
                cell.font = Font(name="Calibri", bold=True, size=10, color="922B21")
                cell.fill = hex_fill("FADBD8")
            elif col_idx == 4 and val > 0:
                cell.font = Font(name="Calibri", bold=True, size=10, color="1E8449")
                cell.fill = hex_fill("D5F5E3")
            elif col_idx == 8 and val > 0:
                cell.font = Font(name="Calibri", bold=True, size=10, color="6C3483")
                cell.fill = hex_fill("F5EEF8")
            else:
                cell.font = Font(name="Calibri", size=10, color="2C3E50")
                cell.fill = row_fill

    # Totals row
    t_row = len(MODULES) + 5
    ws.row_dimensions[t_row].height = 24
    t_values = ["", "GRAND TOTAL", grand_total,
                totals["Happy"], totals["Positive"],
                totals["Negative"], totals["Edge"],
                totals["Security"]]
    for col_idx, val in enumerate(t_values, start=1):
        cell = ws.cell(row=t_row, column=col_idx)
        cell.value = val
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = hex_fill("1F3864")
        cell.border = medium_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Legend
    legend_row = t_row + 2
    ws.merge_cells(f"A{legend_row}:H{legend_row}")
    leg = ws[f"A{legend_row}"]
    leg.value = "Legend"
    leg.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    leg.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[legend_row].height = 20

    for i, (ttype, tstyle) in enumerate(TYPE_STYLE.items()):
        lr = legend_row + 1 + (i // 2)
        lc = 1 + (i % 2) * 3
        ws.merge_cells(start_row=lr, start_column=lc, end_row=lr, end_column=lc + 2)
        lc_cell = ws.cell(row=lr, column=lc)
        lc_cell.value = f"  {tstyle['label']}"
        lc_cell.font = Font(name="Calibri", size=10, color=tstyle["font_color"], bold=True)
        lc_cell.fill = hex_fill(tstyle["fill"])
        lc_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[lr].height = 20


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    build_cover(wb)
    build_summary(wb)

    for module in MODULES:
        build_module_sheet(wb, module)

    # Re-order: Cover first (already is), Summary second (already is)
    out_path = os.path.join(
        r"c:\Vivek_Jain_Custom_changes_Flitesports_Framework_Automation_Playwriter_Python",
        "FliteSports_Bundle_TestCases.xlsx"
    )
    wb.save(out_path)
    print(f"\n✅  File saved: {out_path}")
    total = sum(len(m["cases"]) for m in MODULES)
    print(f"   Modules : {len(MODULES)}")
    print(f"   Total TCs: {total}")


if __name__ == "__main__":
    main()
